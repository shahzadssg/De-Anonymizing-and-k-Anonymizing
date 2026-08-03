#!/usr/bin/env python3
"""
Experimental harness for:
  "De-Anonymizing and k-Anonymizing Individuals using Subject-Object Knowledge Relations"

Every number in Section 6 of the manuscript is produced by this file.

Run from a terminal (this is the canonical, citable invocation):
    python3 harness.py --data test.xlsx --out results/

Run from a Jupyter notebook (for inspection only, not for reported numbers):
    %run harness.py --data test.xlsx --out results
or, if you have pasted this file into a cell:
    main(["--data", "test.xlsx", "--out", "results"])

Design notes
------------
* Graphs are held as dict {subject -> set(object)}. No NetworkX rebuilds inside loops.
  kanon_homogenize() is behaviourally identical to the original notebook implementation
  (verified by --selftest) but does not reconstruct the graph object per user per iteration.
* Utility uses Jensen-Shannon distance, which is bounded in [0,1]. The previous metric
  1 - min(KL, 1) saturated at 0 whenever KL exceeded one nat and could not distinguish
  "degraded" from "destroyed".
* k_degree_anonymize() is the standard Liu & Terzi grouping (sort descending, DP-partition
  into blocks of size k..2k-1, raise every member to the block maximum). The earlier
  implementation was non-monotone in k.
"""

import argparse, itertools, json, os, random, sys, time
from collections import defaultdict

import numpy as np
import pandas as pd
from openpyxl import load_workbook

EPS = 1e-12


# ----------------------------------------------------------------------------- metrics
def js_utility(deg_before, deg_after):
    """1 - Jensen-Shannon distance between degree histograms. Bounded in [0, 1]."""
    lo = min(min(deg_before), min(deg_after))
    hi = max(max(deg_before), max(deg_after))
    bins = np.arange(lo, hi + 2)
    p, _ = np.histogram(deg_before, bins=bins)
    q, _ = np.histogram(deg_after, bins=bins)
    p = p / max(p.sum(), 1)
    q = q / max(q.sum(), 1)
    m = 0.5 * (p + q)

    def kl(a, b):
        mask = a > 0
        return float(np.sum(a[mask] * np.log2(a[mask] / np.maximum(b[mask], EPS))))

    jsd = 0.5 * kl(p, m) + 0.5 * kl(q, m)
    return 1.0 - float(np.sqrt(max(jsd, 0.0)))


def legacy_utility(deg_before, deg_after):
    """The original 1 - min(KL, 1) metric. Retained only to document the saturation."""
    lo = min(min(deg_before), min(deg_after))
    hi = max(max(deg_before), max(deg_after))
    bins = np.arange(lo, hi + 2)
    h1, _ = np.histogram(deg_before, bins=bins, density=True)
    h2, _ = np.histogram(deg_after, bins=bins, density=True)
    h1 = h1 + 1e-10
    h2 = h2 + 1e-10
    h1 = h1 / h1.sum()
    h2 = h2 / h2.sum()
    kl = float(np.sum(h1 * np.log(h1 / h2)))
    return 1.0 - min(kl, 1.0)


def mean_jaccard(G):
    """Mean pairwise Jaccard similarity of subject neighbourhoods. The overlap statistic."""
    subs = list(G)
    if len(subs) < 2:
        return 0.0
    vals = []
    for a, b in itertools.combinations(subs, 2):
        u = len(G[a] | G[b])
        vals.append(len(G[a] & G[b]) / u if u else 0.0)
    return float(np.mean(vals))


def n_edges(G):
    return sum(len(v) for v in G.values())


def equivalence_classes(G):
    cls = defaultdict(list)
    for u, nb in G.items():
        cls[frozenset(nb)].append(u)
    return cls


def verify_k_anonymity(G, k):
    return all(len(v) >= k for v in equivalence_classes(G).values())


def verify_strong_unn(G):
    subs = list(G)
    for i, u in enumerate(subs):
        for v in subs[i + 1:]:
            if G[u] <= G[v] or G[v] <= G[u]:
                return False
    return True


# -------------------------------------------------------------------------- algorithms
def kanon_homogenize(G, k, max_iterations=1000, order=None):
    """
    kAnonHomogenize. Smallest-class-first merging, union as the generalised neighbourhood.
    Adds edges only. `order` fixes the subject processing order so that sensitivity to
    tie-breaking can be measured.
    """
    R = {u: set(v) for u, v in G.items()}
    subs = list(order) if order is not None else list(R)
    added = 0
    for _ in range(max_iterations):
        cls = defaultdict(list)
        for u in subs:
            cls[frozenset(R[u])].append(u)
        vulnerable = [g for g in cls.values() if len(g) < k]
        if not vulnerable:
            break
        vulnerable.sort(key=len)
        merge = set(vulnerable[0])
        if len(merge) < k:
            donors = [u for g in vulnerable[1:] for u in g if u not in merge]
            merge.update(donors[:k - len(merge)])
        if len(merge) < k:
            rest = [u for u in subs if u not in merge]
            merge.update(rest[:k - len(merge)])
        if len(merge) < k:
            break
        target = set().union(*(R[u] for u in merge))
        for u in merge:
            miss = target - R[u]
            added += len(miss)
            R[u] |= miss
    return R, added


def kanon_homogenize_overlap(G, k, order=None):
    """
    Overlap-aware variant. Groups subjects so as to minimise the size of the generalised
    neighbourhood rather than merging the smallest classes in arbitrary order.
    """
    R = {u: set(v) for u, v in G.items()}
    pool = list(order) if order is not None else list(R)
    remaining = set(pool)
    groups = []
    while len(remaining) >= k:
        seed = min(remaining, key=lambda u: (len(R[u]), u))
        grp = [seed]
        remaining.discard(seed)
        cur = set(R[seed])
        while len(grp) < k:
            nxt = min(remaining, key=lambda v: (len(cur | R[v]), v))
            cur |= R[nxt]
            grp.append(nxt)
            remaining.discard(nxt)
        groups.append(grp)
    if remaining:
        if groups:
            groups[-1].extend(remaining)
        else:
            groups.append(list(remaining))
    added = 0
    for grp in groups:
        target = set().union(*(R[u] for u in grp))
        for u in grp:
            miss = target - R[u]
            added += len(miss)
            R[u] |= miss
    return R, added


def kanon_degree_lower_bound(G, k):
    """
    Lower bound on edges needed for k-neighbourhood anonymity.
    Any valid group S has |U(S)| >= max_{u in S} deg(u), so its cost is at least
    |S| * max_deg(S) - sum_deg(S). Optimising that relaxation over all partitions into
    blocks of size >= k is an exact DP on the sorted degree sequence. The bound is loose
    because it ignores the requirement that neighbourhoods be identical, not merely equal
    in size.
    """
    d = sorted(len(v) for v in G.values())
    n = len(d)
    pre = np.concatenate([[0], np.cumsum(d)])
    INF = float("inf")
    dp = [INF] * (n + 1)
    dp[0] = 0.0
    for i in range(1, n + 1):
        for j in range(max(0, i - 2 * k + 1), i - k + 1):
            if dp[j] < INF:
                cand = dp[j] + (i - j) * d[i - 1] - (pre[i] - pre[j])
                dp[i] = min(dp[i], cand)
    return int(dp[n]) if dp[n] < INF else 0


def k_degree_anonymize(G, k):
    """
    k-degree anonymity in the style of Liu and Terzi (2008), restricted to edge addition.
    Sort degrees descending, partition into consecutive blocks of size k..2k-1 by DP,
    raise every member of a block to the block maximum. Monotone non-decreasing in k.
    """
    R = {u: set(v) for u, v in G.items()}
    objs = list(set().union(*R.values())) if R else []
    order = sorted(R, key=lambda u: -len(R[u]))
    d = [len(R[u]) for u in order]
    n = len(d)
    INF = float("inf")
    dp = [INF] * (n + 1)
    cut = [-1] * (n + 1)
    dp[0] = 0.0
    for i in range(1, n + 1):
        for j in range(max(0, i - 2 * k + 1), i - k + 1):
            if dp[j] < INF:
                cand = dp[j] + sum(d[j] - d[t] for t in range(j, i))
                if cand < dp[i]:
                    dp[i], cut[i] = cand, j
    if dp[n] == INF:                       # fewer than k subjects: single block
        blocks = [list(range(n))]
    else:
        blocks, i = [], n
        while i > 0:
            j = cut[i]
            blocks.append(list(range(j, i)))
            i = j
        blocks.reverse()
    added = 0
    for blk in blocks:
        tgt = d[blk[0]]
        for idx in blk:
            u = order[idx]
            need = tgt - len(R[u])
            if need <= 0:
                continue
            cands = [o for o in objs if o not in R[u]]
            pick = cands[:need]
            R[u] |= set(pick)
            added += len(pick)
    return R, added


def deanon_min_graph(G, max_iterations=20000, order=None):
    """
    DeAnonMinGraph. Repairs subset violations by attaching the least-connected object that
    breaks containment, and introduces a synthetic object only when no existing object can.
    """
    R = {u: set(v) for u, v in G.items()}
    subs = list(order) if order is not None else list(R)
    objs = set().union(*R.values()) if R else set()
    obj_deg = defaultdict(int)
    for u in R:
        for o in R[u]:
            obj_deg[o] += 1
    added_edges = 0
    added_nodes = 0
    synth = 0
    for _ in range(max_iterations):
        viol = None
        for i, u in enumerate(subs):
            for v in subs[i + 1:]:
                if R[u] <= R[v] or R[v] <= R[u]:
                    viol = (u, v)
                    break
            if viol:
                break
        if viol is None:
            break
        u, v = viol
        if R[u] <= R[v]:
            u, v = (u, v) if R[u] <= R[v] else (v, u)
        low, high = (u, v) if R[u] <= R[v] else (v, u)
        cands = [o for o in objs if o not in R[high]]
        if cands:
            pick = min(cands, key=lambda o: (obj_deg[o], o))
            R[low].add(pick)
            obj_deg[pick] += 1
            added_edges += 1
        else:
            synth += 1
            new_o = f"__SYNTH_{synth}"
            objs.add(new_o)
            R[low].add(new_o)
            obj_deg[new_o] = 1
            added_edges += 1
            added_nodes += 1
    return R, added_edges, added_nodes


# ------------------------------------------------------------------------------ loaders
def load_nextcloud(path):
    """Parses the Nextcloud 'Shared Folders Overview' export into {user -> set(folder)}."""
    wb = load_workbook(path, read_only=True)
    ws = wb["Shared Folders Overview"]
    rows = []
    skip = False
    folder = None
    for row in ws.iter_rows(values_only=True):
        if not row or all(c is None for c in row):
            continue
        if row[0] in ("Access", "Owner", "Full permissions"):
            skip = True
            continue
        if skip:
            skip = False
            continue
        if row[0] and not str(row[0]).startswith("Column"):
            folder = str(row[0]).strip()
            for cell in row[1:]:
                if cell and str(cell).strip():
                    rows.append((folder, str(cell).strip()))
    G = defaultdict(set)
    for f, u in rows:
        G[u].add(f)
    return dict(G)


def synth_overlap(n_subj, n_obj, deg_lo, deg_hi, n_roles, role_frac, rng):
    """
    Synthetic bipartite generator with a tunable structure knob.
    Each subject takes a deterministic prefix of its role's object bundle, so that subjects
    sharing a role share those objects exactly, and fills the remainder of its degree with
    uniformly drawn private objects. role_frac = 0 gives near-independent neighbourhoods,
    role_frac = 1 gives near-identical neighbourhoods within a role.
    """
    pool = [f"O{i}" for i in range(n_obj)]
    bundles = [sorted(rng.sample(pool, min(n_obj, deg_hi))) for _ in range(n_roles)]
    G = {}
    for i in range(n_subj):
        d = rng.randint(deg_lo, deg_hi)
        b = bundles[i % n_roles]
        n_role = min(len(b), int(round(d * role_frac)))
        nb = set(b[:n_role])
        guard = 0
        while len(nb) < d and guard < 20 * n_obj:
            nb.add(rng.choice(pool))
            guard += 1
        G[f"S{i}"] = nb
    return G


def surrogate_from(G, target_jaccard=None, core_size=40, seed=0, tol=1e-4, iters=24):
    """
    Builds a releasable synthetic stand-in for a private graph.

    Matches, by construction: subject count, object count, the exact degree sequence, and
    (by bisection on the shared-core fraction) the mean pairwise Jaccard similarity.

    IMPORTANT: matching those four statistics does NOT reproduce the anonymization costs of
    the source graph. On our Nextcloud data the surrogate needs 140.8% additional edges for
    3-anonymity against the real graph's 122.6%, and 89.9% against 49.7% at k=2. Mean overlap
    predicts cost well *within* a graph family but does not determine it *across* families,
    because higher-order structure (nesting of neighbourhoods, clustering of the shared core)
    also matters. Use this surrogate to exercise the code, not to reproduce Table 5.
    """
    degs = sorted((len(v) for v in G.values()), reverse=True)
    n_obj = len(set().union(*G.values())) if G else 0
    target = mean_jaccard(G) if target_jaccard is None else target_jaccard

    def build(alpha):
        rng = random.Random(seed)
        pool = [f"O{i}" for i in range(n_obj)]
        core, rest = pool[:core_size], pool[core_size:]
        S = {}
        for i, d in enumerate(degs):
            nc = min(len(core), int(round(alpha * d)))
            nb = set(rng.sample(core, nc))
            need = d - len(nb)
            if need > 0 and rest:
                nb |= set(rng.sample(rest, min(need, len(rest))))
            while len(nb) < d:
                nb.add(rng.choice(pool))
            S[f"S{i}"] = nb
        return S

    lo, hi = 0.0, 0.95
    for _ in range(iters):
        mid = (lo + hi) / 2
        if mean_jaccard(build(mid)) < target:
            lo = mid
        else:
            hi = mid
        if hi - lo < tol:
            break
    return build((lo + hi) / 2)


# -------------------------------------------------------------------------- experiments
def boot_ci(vals, n_boot=10000, alpha=0.05, rng=None):
    rng = rng or np.random.default_rng(0)
    a = np.asarray(vals, dtype=float)
    if len(a) < 2:
        return float(a.mean()), float(a.mean()), float(a.mean())
    idx = rng.integers(0, len(a), size=(n_boot, len(a)))
    means = a[idx].mean(axis=1)
    return (float(a.mean()),
            float(np.percentile(means, 100 * alpha / 2)),
            float(np.percentile(means, 100 * (1 - alpha / 2))))


def exp_real(G, ks, out):
    """Table: every method on the real graph, across k."""
    base = n_edges(G)
    deg0 = [len(v) for v in G.values()]
    rec = []
    t = time.perf_counter()
    Rd, ed, nd = deanon_min_graph(G)
    td = time.perf_counter() - t
    rec.append(dict(method="DeAnonMinGraph", k="", edges_added=ed, pct=100 * ed / base,
                    nodes_added=nd, runtime_s=td,
                    utility_js=js_utility(deg0, [len(v) for v in Rd.values()]),
                    utility_legacy=legacy_utility(deg0, [len(v) for v in Rd.values()]),
                    property_verified=verify_strong_unn(Rd)))
    for k in ks:
        for name, fn in (("kAnonHomogenize", kanon_homogenize),
                         ("kAnonHomogenize-Overlap", kanon_homogenize_overlap),
                         ("k-Degree (Liu-Terzi)", k_degree_anonymize)):
            t = time.perf_counter()
            R, e = fn(G, k)
            el = time.perf_counter() - t
            dg = [len(v) for v in R.values()]
            ok = verify_k_anonymity(R, k) if "kAnon" in name else None
            rec.append(dict(method=name, k=k, edges_added=e, pct=100 * e / base,
                            nodes_added=0, runtime_s=el,
                            utility_js=js_utility(deg0, dg),
                            utility_legacy=legacy_utility(deg0, dg),
                            property_verified=ok))
        rec.append(dict(method="Lower bound (degree relaxation)", k=k,
                        edges_added=kanon_degree_lower_bound(G, k),
                        pct=100 * kanon_degree_lower_bound(G, k) / base,
                        nodes_added=0, runtime_s=0.0, utility_js=None,
                        utility_legacy=None, property_verified=None))
    df = pd.DataFrame(rec)
    df.to_csv(os.path.join(out, "real_results.csv"), index=False)
    return df


def exp_ordering(G, k, n_rep, out, seed=0):
    """Sensitivity of each heuristic to subject processing order on the fixed real graph."""
    rng = random.Random(seed)
    subs = list(G)
    rows = []
    for r in range(n_rep):
        o = subs[:]
        rng.shuffle(o)
        _, e1 = kanon_homogenize(G, k, order=o)
        _, e2 = kanon_homogenize_overlap(G, k, order=o)
        rows.append(dict(rep=r, kAnonHomogenize=e1, kAnonHomogenize_Overlap=e2))
    df = pd.DataFrame(rows)
    df.to_csv(os.path.join(out, "ordering_sensitivity.csv"), index=False)
    return df


def exp_overlap(fracs, n_seeds, out, n_subj=40, n_obj=600, deg_lo=5, deg_hi=120,
                n_roles=8, k=3):
    """Cost as a function of measured neighbourhood overlap. The headline relationship."""
    rows = []
    for f in fracs:
        for s in range(n_seeds):
            rng = random.Random(1000 * s + int(f * 100))
            G = synth_overlap(n_subj, n_obj, deg_lo, deg_hi, n_roles, f, rng)
            base = n_edges(G)
            j = mean_jaccard(G)
            _, e1 = kanon_homogenize(G, k)
            _, e2 = kanon_homogenize_overlap(G, k)
            _, e3 = k_degree_anonymize(G, k)
            _, ed, _ = deanon_min_graph(G)
            rows.append(dict(role_frac=f, seed=s, jaccard=j, base_edges=base,
                             kanon_pct=100 * e1 / base, kanon_overlap_pct=100 * e2 / base,
                             kdegree_pct=100 * e3 / base, deanon_pct=100 * ed / base))
    df = pd.DataFrame(rows)
    df.to_csv(os.path.join(out, "overlap_sweep.csv"), index=False)
    return df


def exp_scalability(sizes, n_seeds, out, n_obj=600, k=3, role_frac=0.3):
    """Both algorithms on the SAME graph ensemble, so the comparison is controlled."""
    rows = []
    for n in sizes:
        for s in range(n_seeds):
            rng = random.Random(77 * s + n)
            G = synth_overlap(n, n_obj, 5, 120, 8, role_frac, rng)
            base = n_edges(G)
            t = time.perf_counter(); _, e1 = kanon_homogenize(G, k); t1 = time.perf_counter() - t
            t = time.perf_counter(); _, ed, _ = deanon_min_graph(G); t2 = time.perf_counter() - t
            rows.append(dict(n_subjects=n, seed=s, base_edges=base,
                             kanon_pct=100 * e1 / base, kanon_time=t1,
                             deanon_pct=100 * ed / base, deanon_time=t2))
    df = pd.DataFrame(rows)
    df.to_csv(os.path.join(out, "scalability.csv"), index=False)
    return df


# ------------------------------------------------------------------------------ selftest
def selftest(G, k=3):
    """
    Confirms the optimised kAnonHomogenize matches the original notebook version.

    Requires mod.py, which is the original notebook exported to a script:
        jupyter nbconvert --to script COSE.ipynb && mv COSE.py mod.py
    Returns True on match, False on mismatch, None if mod.py was not importable.
    """
    try:
        import mod  # the original notebook code, exported to mod.py
    except ImportError:
        print("[selftest] SKIPPED: mod.py not found. The equivalence of "
              "kanon_homogenize() with the original notebook implementation was NOT "
              "verified in this run. Export the notebook with "
              "'jupyter nbconvert --to script COSE.ipynb' and rename the output to "
              "mod.py before citing --selftest as evidence.")
        return None
    _, e_ref, _, _ = mod.kAnonHomogenize(G, k)
    _, e_new = kanon_homogenize(G, k, order=list(G))
    print(f"[selftest] original={e_ref}  harness={e_new}  match={e_ref == e_new}")
    return e_ref == e_new


# --------------------------------------------------------------------------- entrypoint
def _in_notebook():
    """True when running inside an IPython or Jupyter kernel rather than a terminal."""
    if "ipykernel_launcher" in os.path.basename(sys.argv[0] or ""):
        return True
    try:
        from IPython import get_ipython
        ip = get_ipython()
        return ip is not None and "IPKernelApp" in ip.config
    except Exception:
        return False


def _resolve_data(path):
    """
    Look for the dataset relative to the working directory first, then relative to this
    file. Notebook kernels frequently start in a different directory than the script.
    """
    if os.path.isfile(path):
        return os.path.abspath(path)
    here = os.path.join(os.path.dirname(os.path.abspath(__file__)), path)
    if os.path.isfile(here):
        return here
    raise FileNotFoundError(
        f"dataset not found: {path!r}\n"
        f"  looked in cwd:        {os.path.abspath(path)}\n"
        f"  looked next to script: {here}\n"
        f"Pass an explicit path with --data.")


def main(argv=None):
    ap = argparse.ArgumentParser(
        prog="harness.py",
        description="Reproduces every number in Section 6 of the manuscript.")
    ap.add_argument("--data", default="test.xlsx",
                    help="path to the Nextcloud workbook (default: test.xlsx)")
    ap.add_argument("--out", default="results",
                    help="output directory for CSVs (default: results)")
    ap.add_argument("--seeds", type=int, default=30,
                    help="repetitions for the ordering and overlap experiments")
    ap.add_argument("--surrogate", action="store_true",
                    help="write a releasable synthetic stand-in for the private dataset")
    ap.add_argument("--selftest", action="store_true",
                    help="check kanon_homogenize() against the original notebook (needs mod.py)")

    # parse_known_args, not parse_args: a Jupyter kernel injects '-f <kernel.json>' into
    # sys.argv, which an exact parser rejects with SystemExit(2).
    a, extra = ap.parse_known_args(argv)
    if extra:
        print(f"[warn] ignoring unrecognised arguments: {extra}")

    os.makedirs(a.out, exist_ok=True)
    data_path = _resolve_data(a.data)
    print(f"[run] data={data_path}")
    print(f"[run] out={os.path.abspath(a.out)}  seeds={a.seeds}")

    G = load_nextcloud(data_path)
    meta = dict(subjects=len(G), objects=len(set().union(*G.values())),
                edges=n_edges(G), mean_jaccard=mean_jaccard(G),
                distinct_neighbourhoods=len(equivalence_classes(G)),
                max_degree=max(len(v) for v in G.values()),
                median_degree=int(np.median([len(v) for v in G.values()])),
                min_degree=min(len(v) for v in G.values()))
    json.dump(meta, open(os.path.join(a.out, "dataset_meta.json"), "w"), indent=2)
    print("dataset:", json.dumps(meta, indent=2))

    if a.selftest:
        selftest(G)

    print("\n[1/4] real dataset ...")
    print(exp_real(G, [2, 3, 5, 10], a.out).to_string(index=False))
    print("\n[2/4] ordering sensitivity ...")
    od = exp_ordering(G, 3, a.seeds, a.out)
    for c in ("kAnonHomogenize", "kAnonHomogenize_Overlap"):
        m, lo, hi = boot_ci(od[c])
        print(f"   {c}: {m:.1f}  95% CI [{lo:.1f}, {hi:.1f}]")
    print("\n[3/4] overlap sweep ...")
    ov = exp_overlap([0.0, 0.15, 0.3, 0.45, 0.6, 0.75, 0.9, 1.0], a.seeds, a.out)
    print(ov.groupby("role_frac")[["jaccard", "kanon_pct", "kanon_overlap_pct",
                                   "kdegree_pct", "deanon_pct"]].mean().round(2).to_string())
    print("\n[4/4] scalability ...")
    sc = exp_scalability([50, 100, 200, 400, 800, 1600], max(5, a.seeds // 3), a.out)
    print(sc.groupby("n_subjects")[["base_edges", "kanon_pct", "kanon_time",
                                    "deanon_pct", "deanon_time"]].mean().round(3).to_string())
    if a.surrogate:
        print("\n[5/5] synthetic surrogate ...")
        S = surrogate_from(G)
        rows = [dict(subject=u, object=o) for u in S for o in sorted(S[u])]
        pd.DataFrame(rows).to_csv(os.path.join(a.out, "surrogate_graph.csv"), index=False)
        smeta = dict(subjects=len(S), objects=len(set().union(*S.values())),
                     edges=n_edges(S), mean_jaccard=mean_jaccard(S),
                     source_subjects=meta["subjects"], source_objects=meta["objects"],
                     source_edges=meta["edges"], source_mean_jaccard=meta["mean_jaccard"])
        json.dump(smeta, open(os.path.join(a.out, "surrogate_meta.json"), "w"), indent=2)
        print(f"   subjects {smeta['subjects']} (source {smeta['source_subjects']}), "
              f"edges {smeta['edges']} (source {smeta['source_edges']}), "
              f"J {smeta['mean_jaccard']:.4f} (source {smeta['source_mean_jaccard']:.4f})")
        print("   NOTE: matches these statistics only; does not reproduce Table 5 costs.")

    print(f"\nwrote CSVs to {os.path.abspath(a.out)}/")
    return a


if __name__ == "__main__":
    # Inside a kernel, sys.argv belongs to ipykernel_launcher, not to this script, so it
    # is discarded and the defaults are used. Call main([...]) explicitly to override.
    main([] if _in_notebook() else None)
